from typing import Optional, Dict, Tuple

from openpyxl.styles import Font, PatternFill
from openpyxl.styles import NamedStyle
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame, ExcelWriter

from mitosheet.excel_utils import get_column_from_column_index

# Object to map the conditional formatting operators to the excel formulas
CONDITION_TO_COMPARISON_FORMULA: Dict[str, str] = {
    'greater': '>',
    'less': '<',
    'number_exactly': '=',
    'number_not_exactly': '<>',
    'greater_than_or_equal': '>=',
    'less_than_or_equal': '<='
}

SPECIAL_FORMULAS = [
    'string_exactly',
    'string_not_exactly',
    'contains',
    'string_contains_case_insensitive',
    'string_does_not_contain',
    'string_starts_with',
    'string_ends_with',
    'boolean_is_true',
    'boolean_is_false',
]

def get_conditional_format_rule(
    filter_condition: str,
    fill: Optional[PatternFill],
    font: Optional[Font],
    filter_value: str,
    cell_range: str
) -> Optional[FormulaRule]:
    # Update the formulas for the string operators
    comparison = CONDITION_TO_COMPARISON_FORMULA.get(filter_condition)
    if comparison is not None:
        formula = [f'{cell_range}{comparison}{filter_value}']
    elif filter_condition == 'contains':
        formula = [f'NOT(ISERROR(FIND("{filter_value}",{cell_range})))']
    elif filter_condition == 'string_contains_case_insensitive':
        formula = [f'NOT(ISERROR(SEARCH("{filter_value}",{cell_range})))']
    elif filter_condition == 'string_does_not_contain':
        formula = [f'ISERROR(SEARCH("{filter_value}",{cell_range}))']
    elif filter_condition == 'string_starts_with':
        formula = [f'LEFT({cell_range},LEN("{filter_value}"))="{filter_value}"']
    elif filter_condition == 'string_ends_with':
        formula = [f'RIGHT({cell_range},LEN("{filter_value}"))="{filter_value}"']
    elif filter_condition == 'boolean_is_true':
        formula = [cell_range]
    elif filter_condition == 'boolean_is_false':
        formula = [f'NOT({cell_range})']
    elif filter_condition == 'string_exactly':
        formula = [f'EXACT({cell_range},"{filter_value}")']
    elif filter_condition == 'string_not_exactly':
        formula = [f'NOT(EXACT({cell_range},"{filter_value}"))']
    else: return None
    return FormulaRule(fill=fill, font=font, formula=formula)

def add_conditional_formats(
    conditional_formats: list,
    sheet: Worksheet,
    df: DataFrame
) -> None:
    for conditional_format in conditional_formats:
        for filter in conditional_format.get('filters', []):
            # Start with the greater than condition
            operator_info = CONDITION_TO_COMPARISON_FORMULA.get(filter['condition'])
            if operator_info is None and filter['condition'] not in SPECIAL_FORMULAS:
                continue

            # Create the conditional formatting color objects
            cond_fill = None
            cond_font = None
            if conditional_format.get('background_color') is not None:
                cond_fill = PatternFill(start_color=conditional_format['background_color'][1:], end_color=conditional_format['background_color'][1:], fill_type='solid')
            if conditional_format.get('font_color') is not None:
                cond_font = Font(color=conditional_format['font_color'][1:])
            if cond_fill is None and cond_font is None:
                continue
            
            # Add the conditional formatting rule to the sheet
            for column_header in conditional_format['columns']:
                column_index = df.columns.tolist().index(column_header)
                column = get_column_from_column_index(column_index)
                cell_range = f'{column}2:{column}{sheet.max_row}'
                column_conditional_rule = get_conditional_format_rule(
                    filter_condition=filter['condition'],
                    fill=cond_fill,
                    font=cond_font,
                    filter_value=f"{filter['value']}",
                    cell_range=cell_range
                )
                sheet.conditional_formatting.add(cell_range, column_conditional_rule)


def add_formatting_to_excel_sheet(
        writer: ExcelWriter,
        sheet_name: str,
        df: DataFrame,
        header_background_color: Optional[str]=None,
        header_font_color: Optional[str]=None,
        even_background_color: Optional[str]=None,
        even_font_color: Optional[str]=None,
        odd_background_color: Optional[str]=None,
        odd_font_color: Optional[str]=None,
        conditional_formats: Optional[list]=None
    ) -> None:
        """
        Adds formatting to the sheet_name, based on the formatting the user
        currently has applied in the frontend. 

        NOTE: this is a Mito Pro feature.
        """
        workbook = writer.book
        sheet = workbook.get_sheet_by_name(sheet_name)
        
        # Add formatting to the header row   
        header_name = f"{sheet_name}_Header"
        header_format = NamedStyle(name=header_name)
        if header_font_color:
            # Remove the # from the color
            header_format.font = Font(color=header_font_color[1:])
        if header_background_color:
            # Remove the # from the color
            header_format.fill = PatternFill(start_color=header_background_color[1:], end_color=header_background_color[1:], fill_type="solid")
        
        # Only add format if there is a header color or background color
        has_header_formatting = header_background_color or header_font_color 
        if has_header_formatting:
            # Add named styles for the header rows to improve performance
            workbook.add_named_style(header_format)

            # Write the formatting to the sheet
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=1, column=col).style = header_name

        # Add formatting to the rows
        even_name = f"{sheet_name}_Even"
        odd_name = f"{sheet_name}_Odd"
        even_format = NamedStyle(name=even_name)
        odd_format = NamedStyle(name=odd_name)

        # Remove the # from the colors and define the formatting objects
        if even_background_color:
            even_format.fill = PatternFill(start_color=even_background_color[1:], end_color=even_background_color[1:], fill_type="solid")
        if even_font_color:
            even_format.font = Font(color=even_font_color[1:])
        if odd_background_color:
            odd_format.fill = PatternFill(start_color=odd_background_color[1:], end_color=odd_background_color[1:], fill_type="solid")
        if odd_font_color:
            odd_format.font = Font(color=odd_font_color[1:])
        
        # Only add format if there is a background color or font color
        has_row_formatting = even_background_color or even_font_color or odd_background_color or odd_font_color
        if has_row_formatting:
            workbook.add_named_style(even_format)
            workbook.add_named_style(odd_format)

            for row in range(2, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    if row % 2 == 0:
                        sheet.cell(row=row, column=col).style = even_name
                    else:
                        sheet.cell(row=row, column=col).style = odd_name

        # Add conditional formatting
        if conditional_formats is not None:
            add_conditional_formats(conditional_formats, sheet, df)